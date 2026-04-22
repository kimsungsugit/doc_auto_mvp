from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fastapi.testclient import TestClient

from app.main import app, storage


SAMPLE_PATH = ROOT / "samples" / "generated" / "sample_19_statement_type.pdf"
REPORT_PATH = ROOT / "reports" / "demo_run_report.json"


def main() -> None:
    client = TestClient(app)

    with SAMPLE_PATH.open("rb") as file_handle:
        upload_response = client.post(
            "/documents/upload",
            files={"file": (SAMPLE_PATH.name, file_handle.read(), "application/pdf")},
            data={"uploaded_by": "codex-demo"},
        )
    upload_response.raise_for_status()
    upload_payload = upload_response.json()
    document_id = upload_payload["document_id"]

    extract_response = client.post(f"/documents/{document_id}/extract")
    extract_response.raise_for_status()
    extract_payload = extract_response.json()

    fields_response = client.get(f"/documents/{document_id}/fields")
    fields_response.raise_for_status()
    fields_payload = fields_response.json()
    field_map = {field["field_name"]: field for field in fields_payload["fields"]}

    review_response = client.post(f"/documents/{document_id}/review")
    review_response.raise_for_status()

    export_response = client.post(f"/documents/{document_id}/export")
    export_response.raise_for_status()
    export_payload = export_response.json()

    workbook_path = storage.export_path(document_id)
    workbook = load_workbook(workbook_path)
    summary_sheet = workbook[workbook.sheetnames[0]]
    item_sheet = workbook[workbook.sheetnames[1]]

    report = {
        "sample_file": SAMPLE_PATH.name,
        "document_id": document_id,
        "uploaded_file_name": upload_payload["original_file_name"],
        "uploaded_extension": upload_payload["original_extension"],
        "extraction_status": extract_payload["extraction_status"],
        "document_type": extract_payload["document_type"],
        "field_count": extract_payload["field_count"],
        "item_count": extract_payload["item_count"],
        "key_fields": {
            "document_type": field_map["document_type"]["value"],
            "issue_date": field_map["issue_date"]["value"],
            "supplier_name": field_map["supplier_name"]["value"],
            "buyer_name": field_map["buyer_name"]["value"],
            "total_amount": field_map["total_amount"]["value"],
            "item_name": field_map["item_name"]["value"],
        },
        "excel_output": {
            "file_name": export_payload["export_file_name"],
            "path": str(workbook_path),
            "summary_cells": {
                "B3": summary_sheet["B3"].value,
                "F3": summary_sheet["F3"].value,
                "B6": summary_sheet["B6"].value,
                "B8": summary_sheet["B8"].value,
                "F14": summary_sheet["F14"].value,
            },
            "item_rows": [
                {
                    "A3": item_sheet["A3"].value,
                    "E3": item_sheet["E3"].value,
                    "F3": item_sheet["F3"].value,
                }
            ],
        },
    }

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(REPORT_PATH)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
