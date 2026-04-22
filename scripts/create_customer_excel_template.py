from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.settings import CUSTOMER_TEMPLATE_PATH, TEMPLATES_DIR
from app.template_mapping import (
    AUDIT_SHEET_NAME,
    DEFAULT_ITEM_HEADERS,
    DEFAULT_SUMMARY_LAYOUT,
    ITEM_SHEET_NAME,
    SUMMARY_CONTEXT_HEADERS,
    SUMMARY_CONTEXT_START_ROW,
    SUMMARY_SHEET_NAME,
    SUMMARY_TITLE_CELL,
    SUMMARY_TITLE_RANGE,
)


def build_template(path: Path) -> None:
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = SUMMARY_SHEET_NAME

    title_fill = PatternFill("solid", fgColor="1F4E78")
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    subtle_fill = PatternFill("solid", fgColor="F3F7FB")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    summary_sheet.merge_cells(SUMMARY_TITLE_RANGE)
    summary_sheet[SUMMARY_TITLE_CELL] = DEFAULT_SUMMARY_LAYOUT[SUMMARY_TITLE_CELL]
    summary_sheet[SUMMARY_TITLE_CELL].fill = title_fill
    summary_sheet[SUMMARY_TITLE_CELL].font = white_font
    summary_sheet[SUMMARY_TITLE_CELL].alignment = center

    for cell, value in DEFAULT_SUMMARY_LAYOUT.items():
        if cell == SUMMARY_TITLE_CELL:
            continue
        summary_sheet[cell] = value
        summary_sheet[cell].fill = label_fill
        summary_sheet[cell].font = bold_font
        summary_sheet[cell].border = border

    for cell in ["B3", "F3", "B6", "F6", "B8", "F8", "B10", "B14", "D14", "F14", "B17", "F17"]:
        summary_sheet[cell].border = border

    for index, value in enumerate(SUMMARY_CONTEXT_HEADERS, start=1):
        cell = summary_sheet.cell(row=SUMMARY_CONTEXT_START_ROW, column=index, value=value)
        cell.fill = label_fill
        cell.font = bold_font
        cell.border = border

    for row in range(SUMMARY_CONTEXT_START_ROW + 1, SUMMARY_CONTEXT_START_ROW + 5):
        for column in range(1, 4):
            cell = summary_sheet.cell(row=row, column=column)
            cell.border = border
            cell.fill = subtle_fill

    for column, width in {"A": 16, "B": 28, "C": 14, "D": 18, "E": 18, "F": 24}.items():
        summary_sheet.column_dimensions[column].width = width

    item_sheet = workbook.create_sheet(ITEM_SHEET_NAME)
    item_sheet["A1"] = "품목내역"
    item_sheet["A1"].fill = title_fill
    item_sheet["A1"].font = white_font
    item_sheet["A1"].alignment = center
    item_sheet.merge_cells("A1:F1")
    for cell, value in DEFAULT_ITEM_HEADERS.items():
        item_sheet[cell] = value
        item_sheet[cell].fill = label_fill
        item_sheet[cell].font = bold_font
        item_sheet[cell].border = border
    for column, width in {"A": 24, "B": 18, "C": 10, "D": 14, "E": 16, "F": 12}.items():
        item_sheet.column_dimensions[column].width = width
    for row in range(3, 18):
        for column in "ABCDEF":
            item_sheet[f"{column}{row}"].border = border

    audit_sheet = workbook.create_sheet(AUDIT_SHEET_NAME)
    for label, cell in {
        "문서ID": "A1",
        "원본파일": "A2",
        "업로더": "A3",
        "문서상태": "A4",
        "재처리횟수": "A5",
        "필드명": "A7",
        "값": "B7",
        "신뢰도": "C7",
        "검증상태": "D7",
        "수정자": "E7",
    }.items():
        audit_sheet[cell] = label
    for column, width in {"A": 24, "B": 30, "C": 12, "D": 14, "E": 18}.items():
        audit_sheet.column_dimensions[column].width = width

    workbook.save(path)


if __name__ == "__main__":
    build_template(CUSTOMER_TEMPLATE_PATH)
    print(CUSTOMER_TEMPLATE_PATH)
