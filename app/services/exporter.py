from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models import DocumentField, DocumentRecord, EditHistoryEntry, InvoiceLineItem, LogEntry
from app.settings import CUSTOMER_TEMPLATE_PATH
from app.template_mapping import (
    AUDIT_SHEET_NAME,
    DEFAULT_SUMMARY_LAYOUT,
    ITEM_COLUMNS,
    ITEM_ROW_START,
    ITEM_SHEET_NAME,
    SUMMARY_CELL_MAP,
    SUMMARY_CONTEXT_HEADERS,
    SUMMARY_CONTEXT_START_ROW,
    SUMMARY_SHEET_NAME,
    SUMMARY_TITLE_CELL,
    SUMMARY_TITLE_RANGE,
    get_item_headers,
    get_item_sheet_title,
    get_summary_context,
    get_summary_layout,
)

KOREAN_FONT = Font(name="맑은 고딕", size=10)
EMPTY_FILL = PatternFill("solid", fgColor="FFF2CC")
EMPTY_FONT = Font(name="맑은 고딕", size=10, italic=True, color="7F6000")
NUMBER_FIELDS_SUMMARY = {"supply_amount", "tax_amount", "total_amount"}
DATE_FIELDS_SUMMARY = {"issue_date"}
NUMBER_FIELDS_ITEM = {"quantity", "unit_price", "supply_amount", "tax_amount"}


def _to_number(value: str) -> float | None:
    if not value:
        return None
    cleaned = value.replace(",", "").replace(" ", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _apply_korean_font(cell) -> None:
    cell.font = KOREAN_FONT


def _mark_empty(cell) -> None:
    cell.value = "(미입력)"
    cell.fill = EMPTY_FILL
    cell.font = EMPTY_FONT


class ExcelExporter:
    def export(
        self,
        record: DocumentRecord,
        output_path: Path,
        *,
        edit_history: list[EditHistoryEntry] | None = None,
        processing_logs: list[LogEntry] | None = None,
    ) -> str:
        workbook = self._load_or_build_template()
        summary_sheet = workbook[SUMMARY_SHEET_NAME]
        item_sheet = workbook[ITEM_SHEET_NAME]
        fields_by_name = record.fields_by_name
        doc_field = fields_by_name.get("document_type")
        document_type = doc_field.value if doc_field else ""

        self._write_summary_layout(summary_sheet, document_type)
        self._write_summary_values(summary_sheet, fields_by_name)
        self._write_summary_context(summary_sheet, document_type, fields_by_name)
        self._write_item_headers(item_sheet, document_type)
        self._write_item_sheet(item_sheet, record, fields_by_name)
        self._write_audit_sheet(workbook, record, edit_history, processing_logs)
        workbook.save(output_path)
        return output_path.name

    def _write_summary_layout(self, sheet, document_type: str) -> None:
        layout = get_summary_layout(document_type)
        sheet[SUMMARY_TITLE_CELL] = layout[SUMMARY_TITLE_CELL]
        for cell, default_value in DEFAULT_SUMMARY_LAYOUT.items():
            if cell == SUMMARY_TITLE_CELL:
                continue
            sheet[cell] = layout.get(cell, default_value)

    def _write_summary_values(self, sheet, fields_by_name: dict[str, DocumentField]) -> None:
        for field_name, cell_ref in SUMMARY_CELL_MAP.items():
            field = fields_by_name.get(field_name)
            cell = sheet[cell_ref]
            raw_value = field.value if field else ""

            if field_name in NUMBER_FIELDS_SUMMARY:
                number = _to_number(raw_value)
                if number is None:
                    _mark_empty(cell)
                else:
                    cell.value = number
                    cell.number_format = "#,##0"
                    _apply_korean_font(cell)
            elif field_name in DATE_FIELDS_SUMMARY:
                parsed = _to_date(raw_value)
                if parsed is None:
                    if raw_value:
                        cell.value = raw_value
                        _apply_korean_font(cell)
                    else:
                        _mark_empty(cell)
                else:
                    cell.value = parsed
                    cell.number_format = "yyyy-mm-dd"
                    _apply_korean_font(cell)
            elif field_name == "confidence_score":
                number = _to_number(raw_value)
                if number is None and field is not None:
                    number = float(field.confidence)
                if number is None:
                    _mark_empty(cell)
                else:
                    cell.value = number
                    cell.number_format = "0.00"
                    _apply_korean_font(cell)
            else:
                if raw_value:
                    cell.value = raw_value
                    _apply_korean_font(cell)
                else:
                    _mark_empty(cell)

    def _write_summary_context(self, sheet, document_type: str, fields_by_name: dict[str, DocumentField]) -> None:
        self._reset_sheet(sheet, f"A{SUMMARY_CONTEXT_START_ROW}:F{SUMMARY_CONTEXT_START_ROW + 6}")
        header_row = SUMMARY_CONTEXT_START_ROW
        for index, value in enumerate(SUMMARY_CONTEXT_HEADERS, start=1):
            sheet.cell(row=header_row, column=index, value=value)

        for row_offset, (label, value, note) in enumerate(get_summary_context(document_type), start=1):
            row_number = header_row + row_offset
            sheet.cell(row=row_number, column=1, value=label)
            sheet.cell(row=row_number, column=2, value=value)
            sheet.cell(row=row_number, column=3, value=note)

        sheet.cell(row=header_row + 4, column=1, value="대표 품목")
        item_field = fields_by_name.get("item_name")
        sheet.cell(row=header_row + 4, column=2, value=item_field.value if item_field else "")
        sheet.cell(row=header_row + 4, column=3, value="장문 항목은 검수 화면에서 줄바꿈 포함 확인")

    def _write_item_headers(self, sheet, document_type: str) -> None:
        sheet["A1"] = get_item_sheet_title(document_type)
        for cell, value in get_item_headers(document_type).items():
            sheet[cell] = value

    def _write_item_sheet(self, sheet, record: DocumentRecord, fields_by_name: dict[str, DocumentField]) -> None:
        item_f = fields_by_name.get("item_name")
        supply_f = fields_by_name.get("supply_amount")
        tax_f = fields_by_name.get("tax_amount")
        items = record.items or [
            InvoiceLineItem(
                item_name=item_f.value if item_f else "",
                supply_amount=supply_f.value if supply_f else "",
                tax_amount=tax_f.value if tax_f else "",
            )
        ]

        reset_rows = max(38, len(items) + 5)
        self._reset_sheet(sheet, f"A{ITEM_ROW_START}:F{ITEM_ROW_START + reset_rows - 1}")

        thin = Side(style="thin", color="A6A6A6")
        row_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_offset, item in enumerate(items):
            row_number = ITEM_ROW_START + row_offset
            for field_name, column in ITEM_COLUMNS.items():
                cell = sheet[f"{column}{row_number}"]
                raw_value = str(getattr(item, field_name, "") or "")

                if field_name in NUMBER_FIELDS_ITEM:
                    number = _to_number(raw_value)
                    if number is not None:
                        cell.value = number
                        cell.number_format = "#,##0" if field_name != "quantity" else "#,##0.##"
                elif raw_value:
                    cell.value = raw_value
                _apply_korean_font(cell)
                cell.border = row_border

        if items:
            total_row = ITEM_ROW_START + len(items)
            last_data_row = total_row - 1
            label_cell = sheet[f"D{total_row}"]
            label_cell.value = "합계"
            label_cell.font = Font(name="맑은 고딕", size=10, bold=True)
            label_cell.alignment = Alignment(horizontal="right")
            for column in ("E", "F"):
                cell = sheet[f"{column}{total_row}"]
                cell.value = f"=SUM({column}{ITEM_ROW_START}:{column}{last_data_row})"
                cell.number_format = "#,##0"
                cell.font = Font(name="맑은 고딕", size=10, bold=True)
                cell.border = row_border

    def _write_audit_sheet(
        self,
        workbook,
        record: DocumentRecord,
        edit_history: list[EditHistoryEntry] | None = None,
        processing_logs: list[LogEntry] | None = None,
    ) -> None:
        audit_sheet = workbook[AUDIT_SHEET_NAME] if AUDIT_SHEET_NAME in workbook.sheetnames else workbook.create_sheet(AUDIT_SHEET_NAME)
        reset_until = 8 + len(record.fields) + 8 + len(edit_history or []) + 8 + len(processing_logs or []) + 10
        self._reset_sheet(audit_sheet, f"A1:F{max(120, reset_until)}")
        audit_sheet["A1"] = "문서ID"
        audit_sheet["B1"] = record.document_id
        audit_sheet["A2"] = "원본파일"
        audit_sheet["B2"] = record.original_file_name
        audit_sheet["A3"] = "업로더"
        audit_sheet["B3"] = record.uploaded_by
        audit_sheet["A4"] = "문서상태"
        audit_sheet["B4"] = record.status.value
        audit_sheet["A5"] = "재처리횟수"
        audit_sheet["B5"] = record.retry_count
        audit_sheet["A7"] = "필드명"
        audit_sheet["B7"] = "값"
        audit_sheet["C7"] = "신뢰도"
        audit_sheet["D7"] = "검증상태"
        audit_sheet["E7"] = "수정자"
        for row_index, field in enumerate(record.fields, start=8):
            audit_sheet[f"A{row_index}"] = field.label
            audit_sheet[f"B{row_index}"] = field.value
            conf_cell = audit_sheet[f"C{row_index}"]
            conf_cell.value = float(field.confidence)
            conf_cell.number_format = "0.00"
            audit_sheet[f"D{row_index}"] = field.validation_status.value
            audit_sheet[f"E{row_index}"] = field.updated_by or ""
            for col in ("A", "B", "C", "D", "E"):
                _apply_korean_font(audit_sheet[f"{col}{row_index}"])

        next_row = 8 + len(record.fields) + 2
        if edit_history:
            audit_sheet[f"A{next_row}"] = "편집 이력"
            _apply_korean_font(audit_sheet[f"A{next_row}"])
            audit_sheet[f"A{next_row}"].font = Font(name="맑은 고딕", size=10, bold=True)
            header_row = next_row + 1
            for col, label in zip(("A", "B", "C", "D", "E", "F"), ("시각", "필드", "이전값", "새값", "수정자", "사유")):
                audit_sheet[f"{col}{header_row}"] = label
                _apply_korean_font(audit_sheet[f"{col}{header_row}"])
            for offset, entry in enumerate(edit_history, start=1):
                r = header_row + offset
                audit_sheet[f"A{r}"] = entry.timestamp.strftime("%Y-%m-%d %H:%M:%S") if entry.timestamp else ""
                audit_sheet[f"B{r}"] = entry.field_name
                audit_sheet[f"C{r}"] = entry.old_value
                audit_sheet[f"D{r}"] = entry.new_value
                audit_sheet[f"E{r}"] = entry.updated_by or ""
                audit_sheet[f"F{r}"] = entry.comment or ""
                for col in ("A", "B", "C", "D", "E", "F"):
                    _apply_korean_font(audit_sheet[f"{col}{r}"])
            next_row = header_row + len(edit_history) + 2

        if processing_logs:
            audit_sheet[f"A{next_row}"] = "처리 로그"
            audit_sheet[f"A{next_row}"].font = Font(name="맑은 고딕", size=10, bold=True)
            header_row = next_row + 1
            for col, label in zip(("A", "B", "C"), ("시각", "레벨", "메시지")):
                audit_sheet[f"{col}{header_row}"] = label
                _apply_korean_font(audit_sheet[f"{col}{header_row}"])
            for offset, log in enumerate(processing_logs, start=1):
                r = header_row + offset
                audit_sheet[f"A{r}"] = log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else ""
                audit_sheet[f"B{r}"] = log.level
                audit_sheet[f"C{r}"] = log.message
                for col in ("A", "B", "C"):
                    _apply_korean_font(audit_sheet[f"{col}{r}"])

    def _load_or_build_template(self):
        if CUSTOMER_TEMPLATE_PATH.exists():
            workbook = load_workbook(CUSTOMER_TEMPLATE_PATH)
            required_sheets = {SUMMARY_SHEET_NAME, ITEM_SHEET_NAME, AUDIT_SHEET_NAME}
            if required_sheets.issubset(set(workbook.sheetnames)):
                return workbook

        workbook = Workbook()
        summary_sheet = workbook.active or workbook.create_sheet(SUMMARY_SHEET_NAME)
        summary_sheet.title = SUMMARY_SHEET_NAME
        self._build_template(summary_sheet)
        workbook.create_sheet(ITEM_SHEET_NAME)
        workbook.create_sheet(AUDIT_SHEET_NAME)
        return workbook

    def _reset_sheet(self, sheet, cell_range: str) -> None:
        for row in sheet[cell_range]:
            for cell in row:
                cell.value = None

    def _build_template(self, sheet) -> None:
        title_fill = PatternFill("solid", fgColor="1F4E78")
        label_fill = PatternFill("solid", fgColor="D9EAF7")
        subtle_fill = PatternFill("solid", fgColor="F3F7FB")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        bold_font = Font(bold=True)
        thin = Side(style="thin", color="A6A6A6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells(SUMMARY_TITLE_RANGE)
        sheet[SUMMARY_TITLE_CELL] = DEFAULT_SUMMARY_LAYOUT[SUMMARY_TITLE_CELL]
        sheet[SUMMARY_TITLE_CELL].fill = title_fill
        sheet[SUMMARY_TITLE_CELL].font = white_font
        sheet[SUMMARY_TITLE_CELL].alignment = center

        for cell, value in DEFAULT_SUMMARY_LAYOUT.items():
            if cell == SUMMARY_TITLE_CELL:
                continue
            sheet[cell] = value
            sheet[cell].fill = label_fill
            sheet[cell].font = bold_font
            sheet[cell].border = border

        for cell in ["B3", "F3", "B6", "F6", "B8", "F8", "B10", "B14", "D14", "F14", "B17", "F17"]:
            sheet[cell].border = border

        for column, width in {"A": 16, "B": 28, "C": 14, "D": 18, "E": 18, "F": 24}.items():
            sheet.column_dimensions[column].width = width

        for row in range(SUMMARY_CONTEXT_START_ROW, SUMMARY_CONTEXT_START_ROW + 5):
            for column in range(1, 4):
                cell = sheet.cell(row=row, column=column)
                cell.border = border
                if row == SUMMARY_CONTEXT_START_ROW:
                    cell.fill = label_fill
                    cell.font = bold_font
                else:
                    cell.fill = subtle_fill

