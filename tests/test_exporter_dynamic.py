from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import load_workbook

from app.models import (
    DocumentField,
    DocumentRecord,
    DocumentStatus,
    EditHistoryEntry,
    InvoiceLineItem,
    LogEntry,
    ValidationStatus,
)
from app.services.exporter import ExcelExporter
from app.template_mapping import AUDIT_SHEET_NAME, ITEM_ROW_START, ITEM_SHEET_NAME


def _record_with(items_count: int) -> DocumentRecord:
    items = [
        InvoiceLineItem(
            item_name=f"품목 {i+1}",
            quantity="1",
            unit_price="1,000",
            supply_amount="1,000",
            tax_amount="100",
        )
        for i in range(items_count)
    ]
    return DocumentRecord(
        document_id="dyn1",
        original_file_name="dyn.pdf",
        uploaded_by="tester",
        uploaded_at=datetime.now(UTC),
        status=DocumentStatus.REVIEWED,
        file_size=1,
        items=items,
        fields=[
            DocumentField(field_name="document_type", label="유형", value="전자세금계산서", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supply_amount", label="공급가액", value=f"{items_count * 1000:,}", validation_status=ValidationStatus.OK),
        ],
    )


class TestDynamicItems:
    def test_fifty_items_all_written_with_formatting(self, tmp_path):
        out = tmp_path / "fifty.xlsx"
        ExcelExporter().export(_record_with(50), out)
        wb = load_workbook(out)
        items = wb[ITEM_SHEET_NAME]

        last_data_row = ITEM_ROW_START + 49
        assert items[f"A{last_data_row}"].value == "품목 50"
        assert items[f"E{last_data_row}"].value == 1000.0
        assert items[f"A{last_data_row}"].font.name == "맑은 고딕"

    def test_sum_formula_present_after_items(self, tmp_path):
        out = tmp_path / "sum.xlsx"
        ExcelExporter().export(_record_with(3), out)
        wb = load_workbook(out)
        items = wb[ITEM_SHEET_NAME]
        total_row = ITEM_ROW_START + 3

        e_formula = items[f"E{total_row}"].value
        f_formula = items[f"F{total_row}"].value
        assert isinstance(e_formula, str) and e_formula.startswith("=SUM(E")
        assert isinstance(f_formula, str) and f_formula.startswith("=SUM(F")


class TestAuditHistory:
    def test_edit_history_and_logs_rendered(self, tmp_path):
        out = tmp_path / "audit.xlsx"
        record = _record_with(2)
        edit_history = [
            EditHistoryEntry(
                timestamp=datetime(2026, 4, 10, 9, 30, tzinfo=UTC),
                field_name="supply_amount",
                old_value="100",
                new_value="1,000",
                updated_by="reviewer",
                comment="오타 수정",
            ),
        ]
        logs = [
            LogEntry(timestamp=datetime(2026, 4, 10, 9, 0, tzinfo=UTC), level="info", message="업로드 완료"),
            LogEntry(timestamp=datetime(2026, 4, 10, 9, 15, tzinfo=UTC), level="warning", message="OCR 경고"),
        ]
        ExcelExporter().export(record, out, edit_history=edit_history, processing_logs=logs)

        wb = load_workbook(out)
        audit = wb[AUDIT_SHEET_NAME]

        cells_text = [str(c.value) for row in audit.iter_rows() for c in row if c.value]
        assert "편집 이력" in cells_text
        assert "처리 로그" in cells_text
        assert "supply_amount" in cells_text
        assert "오타 수정" in cells_text
        assert "OCR 경고" in cells_text
