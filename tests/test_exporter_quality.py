from __future__ import annotations

from datetime import UTC, datetime

from openpyxl import load_workbook

from app.models import DocumentField, DocumentRecord, DocumentStatus, InvoiceLineItem, ValidationStatus
from app.services.exporter import ExcelExporter
from app.template_mapping import AUDIT_SHEET_NAME, ITEM_SHEET_NAME, SUMMARY_SHEET_NAME


def _record() -> DocumentRecord:
    return DocumentRecord(
        document_id="q1",
        original_file_name="q.pdf",
        uploaded_by="tester",
        uploaded_at=datetime.now(UTC),
        status=DocumentStatus.REVIEWED,
        file_size=1,
        items=[
            InvoiceLineItem(item_name="A", quantity="2", unit_price="10,000", supply_amount="20,000", tax_amount="2,000"),
        ],
        fields=[
            DocumentField(field_name="document_type", label="유형", value="전자세금계산서", validation_status=ValidationStatus.OK),
            DocumentField(field_name="issue_date", label="작성일자", value="2026-04-10", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supply_amount", label="공급가액", value="20,000", validation_status=ValidationStatus.OK, confidence=0.93),
            DocumentField(field_name="tax_amount", label="세액", value="2,000", validation_status=ValidationStatus.OK, confidence=0.9),
            DocumentField(field_name="total_amount", label="합계", value="22,000", validation_status=ValidationStatus.OK),
            DocumentField(field_name="confidence_score", label="신뢰도", value="0.91", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supplier_name", label="공급자", value="", validation_status=ValidationStatus.MISSING),
        ],
    )


class TestExporterQuality:
    def test_amounts_written_as_numbers(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        summary = wb[SUMMARY_SHEET_NAME]

        for cell_ref in ("B14", "D14", "F14"):
            cell = summary[cell_ref]
            assert cell.data_type == "n", f"{cell_ref} not numeric"
            assert cell.number_format == "#,##0"

    def test_issue_date_is_datetime_cell(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        cell = wb[SUMMARY_SHEET_NAME]["F3"]
        assert isinstance(cell.value, datetime)
        assert cell.number_format == "yyyy-mm-dd"

    def test_korean_font_applied(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        summary = wb[SUMMARY_SHEET_NAME]
        assert summary["B14"].font.name == "맑은 고딕"

    def test_empty_fields_marked_as_placeholder(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        summary = wb[SUMMARY_SHEET_NAME]
        supplier_cell = summary["B6"]
        assert supplier_cell.value == "(미입력)"

    def test_confidence_as_decimal(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        audit = wb[AUDIT_SHEET_NAME]
        c8 = audit["C8"]
        assert c8.data_type == "n"
        assert c8.number_format == "0.00"

    def test_item_amounts_numeric(self, tmp_path):
        out = tmp_path / "q.xlsx"
        ExcelExporter().export(_record(), out)
        wb = load_workbook(out)
        items = wb[ITEM_SHEET_NAME]
        assert items["E3"].value == 20000.0
        assert items["E3"].data_type == "n"
        assert items["F3"].value == 2000.0
