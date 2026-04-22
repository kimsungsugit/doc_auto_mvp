from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models import DocumentField, DocumentRecord, DocumentStatus, InvoiceLineItem, ValidationStatus
from app.settings import CUSTOMER_TEMPLATE_PATH
from app.services.exporter import ExcelExporter
from app.template_mapping import AUDIT_SHEET_NAME, ITEM_SHEET_NAME, SUMMARY_SHEET_NAME


def build_record(document_type: str) -> DocumentRecord:
    return DocumentRecord(
        document_id="doc123",
        original_file_name="invoice.pdf",
        uploaded_by="tester",
        uploaded_at=datetime.now(UTC),
        status=DocumentStatus.REVIEWED,
        file_size=10,
        retry_count=2,
        items=[
            InvoiceLineItem(
                item_name="통합 자문 서비스",
                item_spec="분기",
                quantity="1",
                unit_price="150,000",
                supply_amount="150,000",
                tax_amount="15,000",
            ),
            InvoiceLineItem(
                item_name="보안 점검 패키지",
                item_spec="분기",
                quantity="1",
                unit_price="50,000",
                supply_amount="50,000",
                tax_amount="5,000",
            ),
        ],
        fields=[
            DocumentField(field_name="document_type", label="문서 유형", value=document_type, validation_status=ValidationStatus.OK),
            DocumentField(field_name="issue_date", label="작성일자", value="2026-04-10", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supplier_name", label="공급자명", value="테스트랩", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supplier_biz_no", label="공급자 사업자번호", value="119-19-19019", validation_status=ValidationStatus.OK),
            DocumentField(field_name="buyer_name", label="공급받는자명", value="비전상사", validation_status=ValidationStatus.OK),
            DocumentField(field_name="buyer_biz_no", label="공급받는자 사업자번호", value="220-81-10001", validation_status=ValidationStatus.OK),
            DocumentField(field_name="remark", label="비고", value="검수 완료", validation_status=ValidationStatus.OK),
            DocumentField(field_name="item_name", label="품목", value="통합 자문 서비스", validation_status=ValidationStatus.OK),
            DocumentField(field_name="supply_amount", label="공급가액", value="150,000", validation_status=ValidationStatus.OK),
            DocumentField(field_name="tax_amount", label="세액", value="15,000", validation_status=ValidationStatus.OK),
            DocumentField(field_name="total_amount", label="합계금액", value="165,000", validation_status=ValidationStatus.OK),
            DocumentField(field_name="approval_status", label="검수상태", value="Reviewed", validation_status=ValidationStatus.OK),
            DocumentField(field_name="confidence_score", label="문서 신뢰도", value="0.95", validation_status=ValidationStatus.OK),
        ],
    )


def test_exporter_writes_common_cells_items_and_audit() -> None:
    output_dir = Path("storage") / "test_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "exporter_test.xlsx"
    ExcelExporter().export(build_record("전자세금계산서"), output)

    workbook = load_workbook(output)
    sheet = workbook[SUMMARY_SHEET_NAME]
    item_sheet = workbook[ITEM_SHEET_NAME]
    audit_sheet = workbook[AUDIT_SHEET_NAME]

    assert sheet["A1"].value == "문서 자동입력 결과"
    assert sheet["B3"].value == "전자세금계산서"
    assert sheet["F3"].value == datetime(2026, 4, 10)
    assert sheet["F3"].number_format == "yyyy-mm-dd"
    assert sheet["B6"].value == "테스트랩"
    assert sheet["B14"].value == 150000.0
    assert sheet["B14"].number_format == "#,##0"
    assert sheet["F14"].value == 165000.0
    assert sheet["A20"].value == "요약 항목"
    assert sheet["A21"].value == "핵심 체크"
    assert item_sheet["A1"].value == "품목내역"
    assert item_sheet["A2"].value == "항목"
    assert item_sheet["A3"].value == "통합 자문 서비스"
    assert item_sheet["E3"].value == 150000.0
    assert item_sheet["A4"].value == "보안 점검 패키지"
    assert audit_sheet["B1"].value == "doc123"
    assert audit_sheet["B5"].value == 2


def test_exporter_uses_quote_specific_labels_headers_and_context() -> None:
    output_dir = Path("storage") / "test_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "quote_export_test.xlsx"
    ExcelExporter().export(build_record("일반견적서"), output)

    workbook = load_workbook(output)
    sheet = workbook[SUMMARY_SHEET_NAME]
    item_sheet = workbook[ITEM_SHEET_NAME]

    assert sheet["A1"].value == "견적서 자동입력 결과"
    assert sheet["E3"].value == "견적일자"
    assert sheet["A6"].value == "발행 업체"
    assert sheet["A8"].value == "수신처"
    assert sheet["A14"].value == "견적금액"
    assert sheet["E14"].value == "총 견적금액"
    assert sheet["A21"].value == "핵심 체크"
    assert sheet["B24"].value == "통합 자문 서비스"
    assert item_sheet["A1"].value == "견적항목"
    assert item_sheet["A2"].value == "견적 항목"
    assert item_sheet["B2"].value == "세부 내용"


def test_customer_template_exists_after_generation() -> None:
    assert CUSTOMER_TEMPLATE_PATH.exists()
