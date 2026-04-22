from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models import DocumentField, DocumentRecord, DocumentStatus, ValidationStatus
from app.services.exporter import ExcelExporter
from app.template_mapping import AUDIT_SHEET_NAME, ITEM_SHEET_NAME, SUMMARY_SHEET_NAME


def _make_record() -> DocumentRecord:
    fields = [
        DocumentField(field_name="document_type", label="문서유형", value="전자세금계산서",
                      confidence=0.95, validation_status=ValidationStatus.OK, required=True),
        DocumentField(field_name="issue_date", label="작성일자", value="2026-04-06",
                      confidence=0.9, validation_status=ValidationStatus.OK, required=True),
        DocumentField(field_name="supplier_name", label="공급자명", value="ABC상사",
                      confidence=0.9, validation_status=ValidationStatus.OK, required=True),
        DocumentField(field_name="supply_amount", label="공급가액", value="100,000",
                      confidence=0.9, validation_status=ValidationStatus.OK, required=True),
        DocumentField(field_name="tax_amount", label="세액", value="10,000",
                      confidence=0.9, validation_status=ValidationStatus.OK, required=True),
        DocumentField(field_name="total_amount", label="합계", value="110,000",
                      confidence=0.9, validation_status=ValidationStatus.OK, required=True),
    ]
    return DocumentRecord(
        document_id="test1234abcd",
        original_file_name="test.pdf",
        original_extension=".pdf",
        uploaded_by="tester",
        uploaded_at=datetime.now(UTC),
        file_size=100,
        fields=fields,
        status=DocumentStatus.REVIEWED,
    )


class TestExporterBuildsTemplateWhenMissing:
    """Cover _build_template and _load_or_build_template branches."""

    def test_builds_workbook_when_template_absent(self, tmp_path, monkeypatch):
        # Point CUSTOMER_TEMPLATE_PATH to non-existent file
        fake_template = tmp_path / "missing_template.xlsx"
        monkeypatch.setattr("app.services.exporter.CUSTOMER_TEMPLATE_PATH", fake_template)

        exporter = ExcelExporter()
        record = _make_record()
        output = tmp_path / "out.xlsx"

        file_name = exporter.export(record, output)
        assert file_name == "out.xlsx"
        assert output.exists()

        workbook = load_workbook(output)
        assert SUMMARY_SHEET_NAME in workbook.sheetnames
        assert ITEM_SHEET_NAME in workbook.sheetnames
        assert AUDIT_SHEET_NAME in workbook.sheetnames

    def test_rebuilds_workbook_when_template_missing_sheets(self, tmp_path, monkeypatch):
        # Create a broken template missing required sheets
        from openpyxl import Workbook
        broken_template = tmp_path / "broken_template.xlsx"
        wb = Workbook()
        wb.active.title = "OnlyOneSheet"
        wb.save(broken_template)

        monkeypatch.setattr("app.services.exporter.CUSTOMER_TEMPLATE_PATH", broken_template)

        exporter = ExcelExporter()
        record = _make_record()
        output = tmp_path / "out2.xlsx"

        exporter.export(record, output)

        workbook = load_workbook(output)
        assert SUMMARY_SHEET_NAME in workbook.sheetnames
        assert ITEM_SHEET_NAME in workbook.sheetnames
        assert AUDIT_SHEET_NAME in workbook.sheetnames
